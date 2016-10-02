from util_data_storage_and_load import *


import openpyxl

data_folder = '/home/jzh/Dropbox/Research/\
Data-driven_estimation_inverse_optimization/INRIX/Raw_data/'

########## extract tmc info for link_1
# load attribute table link_1 data
wb_link_1 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_1.xlsx')

# get sheet name from workbook
sheet_link_1_name = wb_link_1.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_1 data
sheet_link_1 = wb_link_1.get_sheet_by_name(sheet_link_1_name)

tmc_list_link_1 = []

for i in xrange(2, 1 + sheet_link_1.max_row):
    tmc_list_link_1.append(sheet_link_1.cell(row=i, column=2).value.encode('utf-8'))
    
########## extract tmc info for link_2
# load attribute table link_2 data
wb_link_2 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_2.xlsx')

# get sheet name from workbook
sheet_link_2_name = wb_link_2.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_2 data
sheet_link_2 = wb_link_2.get_sheet_by_name(sheet_link_2_name)

tmc_list_link_2 = []

for i in xrange(2, 1 + sheet_link_2.max_row):
    tmc_list_link_2.append(sheet_link_2.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_3
# load attribute table link_3 data
wb_link_3 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_3.xlsx')

# get sheet name from workbook
sheet_link_3_name = wb_link_3.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_3 data
sheet_link_3 = wb_link_3.get_sheet_by_name(sheet_link_3_name)

tmc_list_link_3 = []

for i in xrange(2, 1 + sheet_link_3.max_row):
    tmc_list_link_3.append(sheet_link_3.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_4
# load attribute table link_4 data
wb_link_4 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_4.xlsx')

# get sheet name from workbook
sheet_link_4_name = wb_link_4.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_4 data
sheet_link_4 = wb_link_4.get_sheet_by_name(sheet_link_4_name)

tmc_list_link_4 = []

for i in xrange(2, 1 + sheet_link_4.max_row):
    tmc_list_link_4.append(sheet_link_4.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_5
# load attribute table link_5 data
wb_link_5 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_5.xlsx')

# get sheet name from workbook
sheet_link_5_name = wb_link_5.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_5 data
sheet_link_5 = wb_link_5.get_sheet_by_name(sheet_link_5_name)

tmc_list_link_5 = []

for i in xrange(2, 1 + sheet_link_5.max_row):
    tmc_list_link_5.append(sheet_link_5.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_6
# load attribute table link_6 data
wb_link_6 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_6.xlsx')

# get sheet name from workbook
sheet_link_6_name = wb_link_6.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_6 data
sheet_link_6 = wb_link_6.get_sheet_by_name(sheet_link_6_name)

tmc_list_link_6 = []

for i in xrange(2, 1 + sheet_link_6.max_row):
    tmc_list_link_6.append(sheet_link_6.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_7
# load attribute table link_7 data
wb_link_7 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_7.xlsx')

# get sheet name from workbook
sheet_link_7_name = wb_link_7.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_7 data
sheet_link_7 = wb_link_7.get_sheet_by_name(sheet_link_7_name)

tmc_list_link_7 = []

for i in xrange(2, 1 + sheet_link_7.max_row):
    tmc_list_link_7.append(sheet_link_7.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_8
# load attribute table link_8 data
wb_link_8 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_8.xlsx')

# get sheet name from workbook
sheet_link_8_name = wb_link_8.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_8 data
sheet_link_8 = wb_link_8.get_sheet_by_name(sheet_link_8_name)

tmc_list_link_8 = []

for i in xrange(2, 1 + sheet_link_8.max_row):
    tmc_list_link_8.append(sheet_link_8.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_9
# load attribute table link_9 data
wb_link_9 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_9.xlsx')

# get sheet name from workbook
sheet_link_9_name = wb_link_9.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_9 data
sheet_link_9 = wb_link_9.get_sheet_by_name(sheet_link_9_name)

tmc_list_link_9 = []

for i in xrange(2, 1 + sheet_link_9.max_row):
    tmc_list_link_9.append(sheet_link_9.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_10
# load attribute table link_10 data
wb_link_10 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_10.xlsx')

# get sheet name from workbook
sheet_link_10_name = wb_link_10.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_10 data
sheet_link_10 = wb_link_10.get_sheet_by_name(sheet_link_10_name)

tmc_list_link_10 = []

for i in xrange(2, 1 + sheet_link_10.max_row):
    tmc_list_link_10.append(sheet_link_10.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_11
# load attribute table link_11 data
wb_link_11 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_11.xlsx')

# get sheet name from workbook
sheet_link_11_name = wb_link_11.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_11 data
sheet_link_11 = wb_link_11.get_sheet_by_name(sheet_link_11_name)

tmc_list_link_11 = []

for i in xrange(2, 1 + sheet_link_11.max_row):
    tmc_list_link_11.append(sheet_link_11.cell(row=i, column=2).value.encode('utf-8'))

########## extract tmc info for link_12
# load attribute table link_12 data
wb_link_12 = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_link_12.xlsx')

# get sheet name from workbook
sheet_link_12_name = wb_link_12.sheetnames[0].encode('utf-8')

# get sheet of attribute table link_12 data
sheet_link_12 = wb_link_12.get_sheet_by_name(sheet_link_12_name)

tmc_list_link_12 = []

for i in xrange(2, 1 + sheet_link_12.max_row):
    tmc_list_link_12.append(sheet_link_12.cell(row=i, column=2).value.encode('utf-8'))


zdump([tmc_list_link_1, tmc_list_link_2, tmc_list_link_3, tmc_list_link_4, tmc_list_link_5, \
	tmc_list_link_6, tmc_list_link_7, tmc_list_link_8, tmc_list_link_9, tmc_list_link_10, \
	tmc_list_link_11, tmc_list_link_12], '../temp_files/tmc_list_links.pkz')



