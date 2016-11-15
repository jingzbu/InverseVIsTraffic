from util_data_storage_and_load import *

import openpyxl

data_folder = '/home/jzh/Dropbox/Research/\
Data-driven_estimation_inverse_optimization/INRIX/Raw_data/'

tmc_list_journal_link_dict = {}
for idx in range(130)[1:]:
    ########## extract tmc info for link_1
    # load attribute table link_1 data
    wb_link = openpyxl.load_workbook(data_folder + 'filtered_INRIX_attribute_table_journal_link_%d.xlsx' %(idx))

    # get sheet name from workbook
    sheet_link_name = wb_link.sheetnames[0].encode('utf-8')

    # get sheet of attribute table link_1 data
    sheet_link = wb_link.get_sheet_by_name(sheet_link_name)

    tmc_list_journal_link = []  

    for i in xrange(2, 1 + sheet_link.max_row):
        tmc_list_journal_link.append(sheet_link.cell(row=i, column=2).value.encode('utf-8')) 
    tmc_list_journal_link_dict[str(idx)] = tmc_list_journal_link


zdump(tmc_list_journal_link_dict, '../temp_files/tmc_list_journal_links_dict.pkz')
