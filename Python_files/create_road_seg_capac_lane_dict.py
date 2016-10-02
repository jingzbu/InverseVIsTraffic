#!/usr/bin/env python


__author__ = "Jing Zhang"
__email__ = "jingzbu@gmail.com"
__status__ = "Development"

##### read in raw data

import openpyxl

data_folder = '/home/jzh/Dropbox/Research/\
Data-driven_estimation_inverse_optimization/INRIX/Raw_data/'

# load filtered capacity attribute table raw data
wb_capac = openpyxl.load_workbook(data_folder + 'capacity_attribute_table.xlsx')

# get sheet name from workbook
sheet_capac_name = wb_capac.sheetnames[0].encode('utf-8')

# get sheet of capacity attribute table raw data
sheet_capac = wb_capac.get_sheet_by_name(sheet_capac_name)

row_count = sheet_capac.max_row

RoadSegCapacLane_dict_JSON = {}

for row_idx in range(row_count + 1)[2:]:
    road_invent = str(sheet_capac.cell(row=row_idx, column=26).value)
    length = sheet_capac.cell(row=row_idx, column=2).value
    route_num = str(sheet_capac.cell(row=row_idx, column=8).value)
    # take the period capacity factor into consideration
    AB_AM_capac = (1.0/2.5) * float(sheet_capac.cell(row=row_idx, column=18).value)
    AB_MD_capac = (1.0/4.75) * float(sheet_capac.cell(row=row_idx, column=20).value)
    AB_PM_capac = (1.0/2.5) * float(sheet_capac.cell(row=row_idx, column=22).value)
    AB_NT_capac = (1.0/7.0) * float(sheet_capac.cell(row=row_idx, column=24).value)
    AB_AM_lane = int(sheet_capac.cell(row=row_idx, column=10).value)
    AB_MD_lane = int(sheet_capac.cell(row=row_idx, column=12).value)
    AB_PM_lane = int(sheet_capac.cell(row=row_idx, column=14).value)
    AB_NT_lane = int(sheet_capac.cell(row=row_idx, column=16).value)
    key = road_invent
    data = {'road_invent': road_invent, 'length': length, \
            'route_num': route_num, 'AB_AM_capac': AB_AM_capac, \
           'AB_MD_capac': AB_MD_capac, 'AB_PM_capac': AB_PM_capac, \
           'AB_NT_capac': AB_NT_capac, 'AB_AM_lane': AB_AM_lane, \
           'AB_MD_lane': AB_MD_lane, 'AB_PM_lane': AB_PM_lane, \
           'AB_NT_lane': AB_NT_lane}
    RoadSegCapacLane_dict_JSON[key] = data
        
import json

# Writing JSON data
with open('../temp_files/RoadSegCapacLane_dict_JSON.json', 'w') as json_file:
    json.dump(RoadSegCapacLane_dict_JSON, json_file)
